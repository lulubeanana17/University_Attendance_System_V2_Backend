from django.shortcuts import render
from rest_framework import status
from rest_framework.authentication import SessionAuthentication, TokenAuthentication
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from openpyxl import load_workbook

from attendance.models import Semester, Course
from attendance.serializer import SemesterSerializer, CourseSerializer, UserSerializer, StudentSerializer


# Create your views here.
@api_view(["GET"])
def index(request):
    return Response(
        {
            "data": [
                {"about": "basic index",
                 "url": "../attendance/",
                 "phone": "basic url, it is a root"
                 },
            ]
        }
    )

class Logout(APIView):
    def get(self, request, format=None):
        request.user.auth_token.delete()
        return Response(status=status.HTTP_200_OK)

@api_view(["GET", "POST"])
def semester_list(request):
    if request.method == 'GET':
        semesters = Semester.objects.all()
        serializer = SemesterSerializer(semesters, many=True)
        return Response(serializer.data)
    if request.method == 'POST':
        serializer = SemesterSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors)

@api_view(["GET", "POST"])
def course_list(request):
    if request.method == 'GET':
        courses = Course.objects.all()
        serializer = CourseSerializer(courses, many=True)
        return Response(serializer.data)
    if request.method == 'POST':
        serializer = CourseSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors)

@api_view(["POST"])
def upload_students_from_excel(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        wb = load_workbook(file)
        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]
        students_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            student_info = dict(zip(headers, row))
            students_data.append(student_info)

        for student_info in students_data:
            user_data = {
                "username": student_info['username'],
                "password": "unitec123",
                "first_name": student_info['firstname'],
                "last_name": student_info['lastname'],
            }
            user_serializer = UserSerializer(data=user_data)
            if user_serializer.is_valid():
                user = user_serializer.save()

                student_data = {
                    "dob": student_info['dob'],
                    "studentInfo": user.id,
                }
                student_serializer = StudentSerializer(data=student_data)
                if student_serializer.is_valid():

                    student_serializer.save()

                    user.student = student_serializer.instance
                    user.save()
                else:
                    pass
            else:
                pass

        return Response({"message": "Students uploaded successfully"}, status=status.HTTP_201_CREATED)
    return Response({"message": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)


class UserTypeCheckView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        # Check if the user is an admin
        if user.is_staff:
            return Response({"user_type": "admin"})

        # Check if the user is a lecturer
        if hasattr(user, 'lecturer'):
            return Response({"user_type": "lecturer"})

        # Check if the user is a student
        if hasattr(user, 'student'):
            return Response({"user_type": "student"})

        return Response({"user_type": "unknown"})

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_enrollment_id(request):
    user = request.user
    try:
        enrollment_id = user.student.enrollment.id
        return Response({"enrollment_id": enrollment_id})
    except AttributeError:
        return Response({"message": "User is not a student"}, status=400)
